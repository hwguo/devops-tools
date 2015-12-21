#!/usr/bin/env python
# -*- coding:utf-8 -*-
#
#   Author  :   henry
#   E-mail  :   gzguohongwei@corp.netease.com
#   Date    :   15/12/21 11:43:07
#   Desc    :   
#

import os
import re
import openpyxl
import datetime


EXCEL_DIR = '/Users/henry/Desktop/haha/excels'


class NoColumnException(Exception):
    message = 'No Hostname Column Found'


def update_db(value, date):
    print '%s, %s' % (value, date)


def parse_sheet(work_sheet, start_date):
    flag = 0
    for i in xrange(1, 26):
        c = work_sheet.cell(column=i, row=1)
        if c.value == 'Hostname' or c.value == 'hostname':
            flag = 1
            for j in xrange(1, work_sheet.max_row):
                cc = work_sheet.cell(column=i, row=j)
                if cc.value != None:
                    update_db(cc.value, start_date)
            break
    if flag == 0:
        print '[worksheet:%s has no column (Hostname)], you should manully add to db' % work_sheet.title
        raise NoColumnException


def main():
    try:
        for filename in os.listdir(EXCEL_DIR):
            efile = '/'.join([EXCEL_DIR, str(filename)])
            dates = re.search(".*-(2015\d*)-.*", filename).groups()
            if len(dates) != 1:
                print '[file: %s] parse create date failed.' % filename
                raise
            else:
                start_date = dates[0]
            wb = openpyxl.load_workbook(efile)
            ws = wb.active
            parse_sheet(ws, start_date)
    except OSError as e:
        print e
    except NoColumnException as e:
        print e.message
    except Exception:
        import sys
        sys.exit(-1)


if __name__ == '__main__':
    main()
